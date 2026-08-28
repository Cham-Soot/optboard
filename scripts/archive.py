#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
만기가 지난 월물을 봉인해 백업으로 남긴다. 매일 수집 뒤에 자동으로 한 번 돈다.

만기 지난 월물은 KIS에서 더 이상 조회되지 않는다. 그래서 이 폴더가 유일한 기록이 된다.
봉인본은 두 벌로 남긴다.

  archive/{YYMM}.json   시세 + 그때의 마킹·메모까지 합친 확정본 (앱에서 다시 열 수 있음)
  archive/{YYMM}.xlsx   행사가마다 시트 1개인 엑셀 (앱 없이 그냥 열어 보는 용도)
  archive/INDEX.md      무엇이 언제 봉인됐는지 목록

한 번 봉인된 월물은 다시 건드리지 않는다. 다시 만들려면 --force.
"""
import argparse
import datetime
import json
import os
import shutil

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
ARCH = os.path.join(ROOT, "archive")

RED, BLUE = "FFC8102E", "FF1B4FA8"


def fmt_strike(s):
    s = float(s)
    return ("%g" % s) if s % 1 else str(int(s))


def this_month():
    n = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
    return "%s%02d" % (str(n.year)[2:], n.month)


def load_marks():
    p = os.path.join(DATA, "marks.json")
    if not os.path.exists(p):
        return {}, {}
    try:
        o = json.load(open(p, encoding="utf-8"))
        return o.get("marks", {}), o.get("memos", {})
    except Exception:
        return {}, {}


def write_xlsx(doc, marks, memos, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    except ImportError:
        print("    openpyxl 이 없어 엑셀 봉인은 건너뜁니다 (json 은 남았습니다)")
        return False

    thin_r = Side(style="medium", color=RED)
    thin_b = Side(style="medium", color=BLUE)
    box_r = Border(left=thin_r, right=thin_r, top=thin_r, bottom=thin_r)
    box_b = Border(left=thin_b, right=thin_b, top=thin_b, bottom=thin_b)
    ul_r = Border(bottom=thin_r)
    ul_b = Border(bottom=thin_b)
    font_r = Font(bold=True, color=RED)
    font_b = Font(bold=True, color=BLUE)
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="FFF2EFE7")

    idx = {fmt_strike(r["strike"]) + "|" + r["date"]: r for r in doc["rows"]}
    wb = Workbook()
    wb.remove(wb.active)

    for strike in doc["strikes"]:
        sk = fmt_strike(strike)
        ws = wb.create_sheet(sk)
        head = ["콜 시가", "콜 고가", "콜 저가", "콜 종가",
                "%s / %s" % (sk, doc.get("label", doc["expiry"])),
                "풋 시가", "풋 고가", "풋 저가", "풋 종가", "메모 1", "메모 2"]
        ws.append(head)
        for c in ws[1]:
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal="center")

        for date in doc["dates"]:
            r = idx.get(sk + "|" + date) or {"c": [None] * 4, "p": [None] * 4,
                                             "memo1": "", "memo2": ""}
            k = doc["expiry"] + "|" + sk + "|" + date
            mm = memos.get(k)
            ws.append(r["c"][:4] + [date] + r["p"][:4] +
                      [mm[0] if mm else (r.get("memo1") or ""),
                       mm[1] if mm else (r.get("memo2") or "")])
            row = ws.max_row
            mk = marks.get(k) or {}
            for cid, col in (("cH", 2), ("cL", 3), ("pH", 7), ("pL", 8)):
                shape = mk.get(cid)
                if not shape:
                    continue
                cell = ws.cell(row=row, column=col)
                hi = cid.endswith("H")
                cell.font = font_r if hi else font_b
                cell.border = (box_r if hi else box_b) if shape == "box" else (ul_r if hi else ul_b)

        for col, w in (("A", 9), ("B", 9), ("C", 9), ("D", 9), ("E", 12),
                       ("F", 9), ("G", 9), ("H", 9), ("I", 9), ("J", 26), ("K", 26)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"

    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true", help="이미 봉인된 월물도 다시 만든다")
    ap.add_argument("--expiry", help="특정 월물만 (예: 2607)")
    args = ap.parse_args()

    os.makedirs(ARCH, exist_ok=True)
    marks, memos = load_marks()
    cutoff = this_month()
    sealed = []

    names = sorted(f[:-5] for f in os.listdir(DATA)
                   if len(f) == 9 and f.endswith(".json") and f[:4].isdigit())
    for ex in names:
        if args.expiry and ex != args.expiry:
            continue
        if ex >= cutoff:                      # 아직 살아 있는 월물은 건드리지 않는다
            continue
        out_json = os.path.join(ARCH, ex + ".json")
        if os.path.exists(out_json) and not args.force:
            continue

        doc = json.load(open(os.path.join(DATA, ex + ".json"), encoding="utf-8"))
        pre = ex + "|"
        doc["sealed"] = datetime.datetime.now(
            datetime.timezone(datetime.timedelta(hours=9))).isoformat(timespec="seconds")
        doc["marks"] = {k: v for k, v in marks.items() if k.startswith(pre)}
        doc["memos"] = {k: v for k, v in memos.items() if k.startswith(pre)}
        # 메모는 행에도 심어 둔다 — 봉인본만 따로 열어도 보이도록
        for r in doc["rows"]:
            k = pre + fmt_strike(r["strike"]) + "|" + r["date"]
            if k in doc["memos"]:
                r["memo1"], r["memo2"] = doc["memos"][k][0], doc["memos"][k][1]

        json.dump(doc, open(out_json, "w", encoding="utf-8"),
                  ensure_ascii=False, separators=(",", ":"))
        ok = write_xlsx(doc, marks, memos, os.path.join(ARCH, ex + ".xlsx"))
        sealed.append((ex, doc, ok))
        print("  봉인 %s — 거래일 %d일, 행사가 %d개, 마킹 %d건"
              % (ex, len(doc["dates"]), len(doc["strikes"]), len(doc["marks"])))

    write_index()
    if not sealed:
        print("새로 봉인할 월물이 없습니다.")
    return 0


def write_index():
    rows = []
    for f in sorted(os.listdir(ARCH)):
        if not f.endswith(".json"):
            continue
        d = json.load(open(os.path.join(ARCH, f), encoding="utf-8"))
        rows.append("| %s | %s | %d일 | %s ~ %s | %s | %s |" % (
            d["expiry"], d.get("label", ""), len(d.get("dates", [])),
            (d.get("dates") or [""])[0], (d.get("dates") or [""])[-1],
            "%d건" % len(d.get("marks", {})),
            (d.get("sealed") or "")[:10]))
    body = ["# 봉인된 월물", "",
            "만기가 지난 월물은 증권사 API로 다시 받을 수 없습니다. 이 폴더가 유일한 기록입니다.",
            "`.json` 은 앱의 **데이터 불러오기** 로 열 수 있고, `.xlsx` 는 그냥 엑셀로 열면 됩니다.", "",
            "| 월물 | 이름 | 거래일 | 기간 | 마킹 | 봉인일 |",
            "|---|---|---|---|---|---|"] + rows + [""]
    open(os.path.join(ARCH, "INDEX.md"), "w", encoding="utf-8").write("\n".join(body))


if __name__ == "__main__":
    raise SystemExit(main())
